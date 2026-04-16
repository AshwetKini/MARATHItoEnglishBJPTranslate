"""
================================================================================
  Marathi → English Name Transliteration Tool (Voter Data)
================================================================================
  
  USAGE:
    1. Place your .xlsx files in the 'input' folder
    2. Run:  python translate.py
    3. Find results in the 'output' folder (prefixed with "EN_")

  This script transliterates Marathi (Devanagari) names into English for:
    - "मतदाराचे पूर्ण नाव"  →  "Voter Full Name (English)"
    - "नातेवाईकाचे पूर्ण नाव"  →  "Relative Full Name (English)"

  All other columns and data remain completely untouched.
================================================================================
"""

import os
import re
import sys
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from indic_transliteration import sanscript


# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(SCRIPT_DIR, "input")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")

# ── Source Marathi column names ───────────────────────────────────────────────
VOTER_NAME_COL = "मतदाराचे पूर्ण नाव"
RELATIVE_NAME_COL = "नातेवाईकाचे पूर्ण नाव"

# ── New English column names ──────────────────────────────────────────────────
VOTER_NAME_EN = "Voter Full Name (English)"
RELATIVE_NAME_EN = "Relative Full Name (English)"

# ── IAST vowel characters (used for schwa deletion logic) ────────────────────
IAST_VOWELS = set("aāiīuūeēoōṛṝ")

# ── Marathi letter-name → English letter (for initials like टी. → T.) ─────
# In Marathi, English letter initials are written phonetically in Devanagari.
# e.g. "टी." is how you write "T." — it's the letter name, NOT a syllable.
MARATHI_INITIALS_MAP = {
    'ए': 'A', 'बी': 'B', 'सी': 'C', 'डी': 'D', 'ई': 'E',
    'एफ': 'F', 'जी': 'G', 'एच': 'H', 'आय': 'I', 'जे': 'J',
    'के': 'K', 'एल': 'L', 'एम': 'M', 'एन': 'N', 'ओ': 'O',
    'पी': 'P', 'क्यू': 'Q', 'आर': 'R', 'एस': 'S', 'टी': 'T',
    'यू': 'U', 'व्ही': 'V', 'डब्ल्यू': 'W', 'एक्स': 'X',
    'वाय': 'Y', 'झेड': 'Z',
}


# ==============================================================================
#  Transliteration Engine
# ==============================================================================

def transliterate_marathi(text):
    """
    Convert Marathi (Devanagari) text to clean English transliteration.
    
    Uses a 5-step process:
      1. IAST transliteration (scholarly standard)
      2. Marathi schwa deletion (remove inherent trailing 'a')
      3. Context-aware anusvara handling (ं → m/n based on following consonant)
      4. Diacritics removal (IAST → plain English letters)
      5. Title-case formatting
    """
    if not text or not isinstance(text, str):
        return ""
    
    text = text.strip()
    if not text:
        return ""
    
    # If text has no Devanagari characters, return as-is
    if not any('\u0900' <= ch <= '\u097F' for ch in text):
        return text
    
    # ── Step 0a: Convert Marathi initials (टी. → T., बी. → B., etc.) ─────
    # Detect Devanagari letter-names followed by a period and replace them
    # with the corresponding single English letter.
    # Important: initials can be written compactly like "सी.एस." (no spaces).
    # So we do regex replacements over the whole string, not only by tokens.
    for dev, eng in MARATHI_INITIALS_MAP.items():
        text = re.sub(rf'{re.escape(dev)}\.', f'{eng}.', text)
    
    # ── Step 0b: Replace Marathi-specific vowel signs ─────────────────────
    # These characters (used in Marathi for English loanwords / foreign
    # sounds) are NOT handled by the indic_transliteration library — they
    # pass through as raw Devanagari and get stripped later. Map them to
    # the closest standard Devanagari equivalents before IAST conversion.
    text = text.replace('\u0949', '\u094B')  # ॉ (candra O sign) → ो (O sign)
    text = text.replace('\u0945', '\u0947')  # ॅ (candra E sign) → े (E sign)
    text = text.replace('\u0911', '\u0913')  # ऑ (independent candra O) → ओ
    text = text.replace('\u090D', '\u090F')  # ऍ (independent candra E) → ए
    text = text.replace('\u0972', '\u0905')  # ॲ (candra A) → अ
    text = text.replace('\u0901', '\u0902')  # ँ (chandrabindu) → ं (anusvara)
    
    # After pre-processing, if no Devanagari remains, return as-is
    if not any('\u0900' <= ch <= '\u097F' for ch in text):
        # Still normalize standalone single-letter initials like "C." -> "C"
        # (these can happen when the input contains only initials, e.g. "टी. बी.")
        text = re.sub(r'\b([A-Z])\.([A-Z])\.', r'\1. \2.', text)  # C.S. -> C. S.
        # Remove dot only for truly standalone initials (not part of chains like "C. S.")
        text = re.sub(
            r'(?<!\b[A-Z]\.\s)\b([A-Z])\.(?!\s*[A-Z]\.)',
            r'\1',
            text
        )
        # If a dotted initial is immediately followed by a new capitalized word,
        # add a space: "C. S.Patil" -> "C. S. Patil"
        text = re.sub(r'(\b[A-Z]\.)\s*(?=[A-Z])', r'\1 ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text.title()
    
    # ── Step 1: IAST transliteration ──────────────────────────────────────
    iast = sanscript.transliterate(text, sanscript.DEVANAGARI, sanscript.IAST)
    
    # ── Step 1b: Clean up nukta-produced combining diacritics ─────────────
    # ड़ → r̤ (r + combining dot below), ढ़ → r̤h — normalize to 'r'
    # ख़ → k͟h (k + combining macron below) — normalize to 'kh'
    iast = iast.replace('r\u0324', 'r')   # r + combining diaeresis below → r
    iast = iast.replace('k\u035F', 'k')   # k + combining double macron below → k
    
    # ── Step 2: Marathi schwa deletion ────────────────────────────────────
    # In Marathi, the inherent vowel 'a' (schwa) at the end of a word is
    # typically silent. In IAST, this appears as a lowercase 'a' after a
    # consonant. Long 'ā' (from explicit matra) is NOT deleted.
    #
    # IMPORTANT: We detect true consonant conjuncts (e.g., द्र, त्र) by
    # checking the ORIGINAL Devanagari text for virama (्) before the last
    # consonant. Digraphs in IAST like 'kh', 'gh', 'th' represent SINGLE
    # Devanagari consonants and should NOT prevent schwa deletion.
    
    # Check if original Devanagari word ends with a conjunct (has virama
    # near the end), which means the trailing 'a' in IAST is needed.
    original_words = text.split()
    
    words = iast.split()
    processed_words = []
    for i, word in enumerate(words):
        if (len(word) > 2
                and word[-1] == 'a'
                and word[-2].lower() not in IAST_VOWELS):
            # Check the original Devanagari word for a virama (्) near the end.
            # Virama before the last consonant means it's a conjunct like द्र,
            # and the trailing 'a' is the real vowel of the final consonant.
            has_conjunct_ending = False
            if i < len(original_words):
                orig = original_words[i]
                # Look for virama (\u094D) in last 3 characters of original word
                # (conjuncts like द्र have virama between the two consonants)
                if len(orig) >= 2:
                    # Check if any of the last few chars contain virama
                    tail = orig[-3:] if len(orig) >= 3 else orig
                    # Virama should be present but NOT be the very last char
                    # (if virama is last, the consonant has no vowel - halant form)
                    virama_positions = [j for j, ch in enumerate(tail) if ch == '\u094D']
                    for vp in virama_positions:
                        # Virama followed by another consonant = conjunct ending
                        if vp < len(tail) - 1:
                            has_conjunct_ending = True
            
            if has_conjunct_ending:
                # Consonant cluster ending — keep the 'a' (e.g., chandra, indra, putra)
                pass
            else:
                word = word[:-1]
        processed_words.append(word)
    iast = ' '.join(processed_words)
    
    # ── Step 3: Context-aware anusvara (ṃ) handling ───────────────────────
    # Before labials (p, b, m, bh, ph) → 'm'
    # Before all others → 'n'
    iast = re.sub(r'[ṃṁ](?=[pbmPBM])', 'm', iast)
    iast = re.sub(r'[ṃṁ]', 'n', iast)
    
    # ── Step 3b: Marathi-specific conjunct mappings ───────────────────────
    # jñ (ज्ञ) → "dnya" in Marathi (not "gya" like Hindi)
    iast = iast.replace('jñ', 'dny')
    iast = iast.replace('Jñ', 'Dny')
    
    # ── Step 3c: Convert IAST 'c' → 'ch' for natural English ─────────────
    # In IAST, च = 'c' and छ = 'ch'. In English transliteration of Indian
    # names, both are written with 'ch'. We convert 'c' → 'ch' first,
    # which makes 'ch' (छ) become 'chh' — standard English convention.
    iast = re.sub(r'(?<![s])c(?!h)', 'ch', iast)   # c → ch (avoid sc, already-ch)
    # Uppercase 'C' can appear for our Marathi-English letter initials
    # (e.g. "सी." is preprocessed to "C."). Avoid converting those standalone
    # initials into "Ch." by only applying when the next character is not
    # whitespace/dot punctuation.
    iast = re.sub(r'(?<![S])C(?!h)(?![.\s])', 'Ch', iast)   # C → Ch (inside words)
    
    # ── Step 4: Replace IAST diacritics with plain English ────────────────
    diacritics_map = [
        # Long vowels → short equivalents
        ('ā', 'a'), ('ī', 'i'), ('ū', 'u'),
        # Vocalic R
        ('ṛ', 'ru'), ('ṝ', 'ru'),
        # Visarga
        ('ḥ', 'h'),
        # Nasals
        ('ṅ', 'n'), ('ñ', 'n'), ('ṇ', 'n'),
        # Retroflex consonants
        ('ṭ', 't'), ('ḍ', 'd'),
        # Sibilants
        ('ś', 'sh'), ('ṣ', 'sh'),
        # Marathi-specific lateral
        ('ḷ', 'l'),
    ]
    
    result = iast
    for iast_char, eng_char in diacritics_map:
        result = result.replace(iast_char, eng_char)
        # Handle uppercase variants too
        result = result.replace(iast_char.upper(), eng_char.capitalize())
    
    # ── Step 5: Clean up and title-case ───────────────────────────────────
    # Remove any remaining non-ASCII characters
    result = re.sub(r'[^\x20-\x7E]+', '', result)
    # Collapse multiple spaces
    result = re.sub(r'\s+', ' ', result).strip()
    # Professional title case
    result = result.title()
    
    # If initials are written compactly, add a space between them: "C.S." -> "C. S."
    result = re.sub(r'\b([A-Z])\.([A-Z])\.', r'\1. \2.', result)
    
    # Remove dots only for standalone initials like "C." -> "C",
    # but keep dots when followed by another dotted initial: "C. S." stays "C. S."
    result = re.sub(
        r'(?<!\b[A-Z]\.\s)\b([A-Z])\.(?!\s*[A-Z]\.)',
        r'\1',
        result
    )
    # If a dotted initial is immediately followed by a new capitalized word,
    # add a space: "C. S.Patil" -> "C. S. Patil"
    result = re.sub(r'(\b[A-Z]\.)\s*(?=[A-Z])', r'\1 ', result)

    return result


# ==============================================================================
#  Excel Processing
# ==============================================================================

def find_target_columns(ws, max_search_rows=10):
    """
    Search the first few rows of a worksheet to find the header row
    and the column indices for the target Marathi name columns.
    
    Returns: (header_row, voter_col_index, relative_col_index)
    """
    header_row = None
    voter_col = None
    relative_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=min(max_search_rows, ws.max_row)):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if val == VOTER_NAME_COL:
                voter_col = cell.column
                header_row = cell.row
            elif val == RELATIVE_NAME_COL:
                relative_col = cell.column
                if header_row is None:
                    header_row = cell.row
    
    return header_row, voter_col, relative_col


def style_header_cell(cell):
    """Apply professional styling to a header cell."""
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def process_excel(input_path, output_path):
    """
    Process a single Excel file:
      - Find the target Marathi columns
      - Transliterate each name to English
      - Add new English columns at the end
      - Save to output path
    
    All existing data, formatting, and structure is preserved.
    """
    filename = os.path.basename(input_path)
    print(f"\n  📄 Processing: {filename}")
    
    wb = openpyxl.load_workbook(input_path)
    sheets_processed = 0
    total_translated = 0
    
    for ws in wb.worksheets:
        header_row, voter_col, relative_col = find_target_columns(ws)
        
        if not voter_col and not relative_col:
            print(f"     Sheet '{ws.title}': No target columns found — skipping (data untouched)")
            continue
        
        data_rows = ws.max_row - header_row
        print(f"     Sheet '{ws.title}': Found columns, processing {data_rows} rows...")
        
        # ── Determine positions for new English columns ───────────────────
        next_col = ws.max_column + 1
        
        en_voter_col = None
        en_relative_col = None
        
        if voter_col:
            en_voter_col = next_col
            next_col += 1
        if relative_col:
            en_relative_col = next_col
        
        # ── Write styled headers ─────────────────────────────────────────
        if en_voter_col:
            hdr = ws.cell(row=header_row, column=en_voter_col, value=VOTER_NAME_EN)
            style_header_cell(hdr)
        
        if en_relative_col:
            hdr = ws.cell(row=header_row, column=en_relative_col, value=RELATIVE_NAME_EN)
            style_header_cell(hdr)
        
        # ── Transliterate each row ───────────────────────────────────────
        translated_count = 0
        start_time = time.time()
        
        for row_num in range(header_row + 1, ws.max_row + 1):
            # Voter name
            if en_voter_col and voter_col:
                original = ws.cell(row=row_num, column=voter_col).value
                if original and str(original).strip():
                    english = transliterate_marathi(str(original))
                    ws.cell(row=row_num, column=en_voter_col, value=english)
                    translated_count += 1
            
            # Relative name
            if en_relative_col and relative_col:
                original = ws.cell(row=row_num, column=relative_col).value
                if original and str(original).strip():
                    english = transliterate_marathi(str(original))
                    ws.cell(row=row_num, column=en_relative_col, value=english)
            
            # Progress indicator for large files
            if row_num % 500 == 0:
                elapsed = time.time() - start_time
                print(f"       ... {row_num - header_row}/{data_rows} rows ({elapsed:.1f}s)")
        
        # ── Auto-fit column widths for new columns ────────────────────────
        for col_idx in [en_voter_col, en_relative_col]:
            if col_idx:
                max_width = 30  # minimum width
                for r in range(header_row, min(header_row + 100, ws.max_row + 1)):
                    val = ws.cell(row=r, column=col_idx).value
                    if val:
                        max_width = max(max_width, len(str(val)) + 4)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_width, 45)
        
        elapsed = time.time() - start_time
        print(f"     ✅ Sheet '{ws.title}': {translated_count} names transliterated ({elapsed:.1f}s)")
        sheets_processed += 1
        total_translated += translated_count
    
    # ── Save output ──────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"  💾 Saved → {os.path.basename(output_path)}  ({total_translated} total translations)")
    
    return sheets_processed, total_translated


# ==============================================================================
#  Main Entry Point
# ==============================================================================

def main():
    print()
    print("=" * 65)
    print("   📋 Marathi → English  |  Voter Name Transliteration Tool")
    print("=" * 65)
    
    # ── Ensure directories exist ──────────────────────────────────────────
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # ── Find all Excel files in input folder ──────────────────────────────
    excel_files = sorted([
        f for f in os.listdir(INPUT_DIR)
        if f.lower().endswith(('.xlsx',))
        and not f.startswith('~$')       # skip temp Excel files
    ])
    
    if not excel_files:
        print(f"\n  ⚠️  No .xlsx files found in: {INPUT_DIR}")
        print(f"  📁 Place your Excel files in the 'input' folder and run again.")
        print()
        sys.exit(0)
    
    print(f"\n  📂 Found {len(excel_files)} file(s) in input folder:")
    for f in excel_files:
        size_kb = os.path.getsize(os.path.join(INPUT_DIR, f)) / 1024
        print(f"     • {f}  ({size_kb:.0f} KB)")
    
    # ── Process each file ─────────────────────────────────────────────────
    total_files = len(excel_files)
    success_count = 0
    error_list = []
    grand_total_names = 0
    
    overall_start = time.time()
    
    for idx, filename in enumerate(excel_files, 1):
        input_path = os.path.join(INPUT_DIR, filename)
        
        # Output filename: prefix with "EN_" (avoid double-prefixing)
        out_name = f"EN_{filename}" if not filename.startswith("EN_") else filename
        output_path = os.path.join(OUTPUT_DIR, out_name)
        
        print(f"\n  [{idx}/{total_files}]", end="")
        
        try:
            sheets, names = process_excel(input_path, output_path)
            success_count += 1
            grand_total_names += names
        except Exception as e:
            print(f"\n  ❌ Error processing {filename}: {e}")
            error_list.append((filename, str(e)))
    
    # ── Summary ───────────────────────────────────────────────────────────
    total_time = time.time() - overall_start
    
    print()
    print("=" * 65)
    print(f"   ✅ DONE — {success_count}/{total_files} files processed successfully")
    print(f"   📊 {grand_total_names} names transliterated in {total_time:.1f}s")
    
    if error_list:
        print(f"\n   ⚠️  {len(error_list)} file(s) had errors:")
        for fname, err in error_list:
            print(f"      • {fname}: {err}")
    
    print(f"\n   📁 Output folder: {OUTPUT_DIR}")
    print("=" * 65)
    print()


if __name__ == "__main__":
    main()
